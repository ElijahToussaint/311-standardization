# imports
import os
import pandas
import openpyxl
import datetime
from openpyxl import load_workbook


# function that iterates through the columns (first row) of an excel file
# the function returns an int (column cell value position)
def selectColumn(columnName):
    # create a list to store column values
    placeList = []
    # iterate through the columns (first row) and store there values
    for column in ws.iter_cols(max_row=1, values_only=True):
        for cell in column:
            placeList.append(cell)
    places = placeList
    # iterate through the column values and save the requested column cell
    for place in places:
        if place == columnName:
            result = placeList.index(place)
            break
        else:
            result = None
    try:
        print(placeList[result])
    except:
        print('%s does not exist...' % columnName)
    return result


# function that gets the cell values of an entire column in an excel file
# the function returns a list of strings (cell values)
def getColumn(columnName):
    if selectColumn(columnName) != None:
        # variable that will store the column value
        # +1 is added to the variable because min_col and max_col has a range of cell value +1
        columnValue = selectColumn(columnName) + 1
        # list that will store the cell values of the selected column
        cellList = []
        # iterate through the selected column and stores its cell values
        for column in ws.iter_cols(min_col=columnValue, max_col=columnValue, min_row=2, values_only=True):
            for cell in column:
                if cell != None:
                    cellList.append(cell.strip())
                else:
                    cellList.append(cell)
        # prints to console the cells that are stored in the list
        cells = cellList
        for cell in cells:
            if cell != None:
                print(cell)
        result = cells
    else:
        result = None
        print('%s does not exist...' % columnName)
    return result


# function the matches two columns (cell values) and adds them to a dictionary
# the function returns a list of dictionaries (dictionaries contains strings)
# the dictionary values are stored in this format -> {column1 : column2} -> column1 = column2
def matchColumn(column1, column2):
    # variables that store the list of columns by calling the getColumn() function
    firstColumn = getColumn(column1)
    secondColumn = getColumn(column2)
    if firstColumn == None or secondColumn == None:
        result = None
        print('Columns could not be matched.') 
    else:
        # list that stores the column values (strings) as row values (ints)
        firstColumnValues = []
        secondColumnValues = []
        # list that stores the dictionary values of both columns
        dictsList = []
        # iterates through the first column and store the values as rows (ints)
        for row in firstColumn:
            if row != None:
                firstColumnValue = firstColumn.index(row)
                firstColumnValues.append(firstColumnValue)
        # iterates through the second column and store the values as rows (ints)
        for row in secondColumn:
            if row != None:
                secondColumnValue = secondColumn.index(row)
                secondColumnValues.append(secondColumnValue)
        # stores the row values (ints) as a list
        # converts the lists to sets and merges them (removing duplicate values)
        matchList = list(set(firstColumnValues).intersection(secondColumnValues))
        print(matchList)
        # iterates through the list of row values
        # converts the rows values (ints) back into strings
        # stores the the values as a list of dictionaries (strings)
        for value in matchList:
            print(firstColumn[value] + ' = ' + secondColumn[value])
            dicts = {firstColumn[value]: secondColumn[value]}
            dictsList.append(dicts)
        result = dictsList
    print(result)
    return result


# function that reads and writes a csv file
def createFile(inputFile, columnName, outputFile):
    if selectColumn(columnName) != None:
        try:
            # read the csv file using pandas
            readFile = pandas.read_csv(
                inputFile, delimiter=',', header=0, index_col=False, low_memory=False, encoding='iso-8859-13', chunksize=100000)
            print('File found. Now reading file...')
            # list of dictionaries that will be used as rows for the output file
            masterDicts = matchColumn(columnName, 'Master_List')
            colIdDicts = matchColumn(columnName, 'Col_ID')
            # list of the first row of the input csv file
            firstColumn = []
            # lists that store the Master_List and Col_ID columns
            newMasterColumn = []
            newColIdColumn = []
            # list that stores the partitioned input file data (DataFrame)
            chunk_list = []
            # iterates through the chunks of input file data
            # each chunk is in DataFrame format
            for chunk in readFile:
                chunk_list.append(chunk)
            df_concat = pandas.concat(chunk_list)
            print(df_concat)
            # iterate through the input file data columns and stores them in a list
            for col in df_concat.columns:
                print(col.strip())
                firstColumn.append(col)
            print(firstColumn)
            print(masterDicts)
            print(colIdDicts)
            # iterate through the Master_List column list and create a new column list
            for col in firstColumn:
                result = 'This attribute does not exist... (' + col + ')'
                for dicts in masterDicts:
                    for x, y in dicts.items():
                        if x == col:
                            result = y
                newMasterColumn.append(result)
            print(newMasterColumn)
            # iterate through the Col_ID column list and create a new column list
            for col in firstColumn:
                result = '9999'
                for dicts in colIdDicts:
                    for x, y in dicts.items():
                        if x == col:
                            result = y
                newColIdColumn.append(result)
            print(newColIdColumn)
            # create a new DataFrame and adding the Master_List as the defualt column
            df = pandas.DataFrame(columns=newMasterColumn)
            # add the Col_ID row as a new row to the new DataFrame
            df.loc[0] = newColIdColumn
            # replace the old column with the new Master_List column
            # this is important when appending Dataframes
            df_concat.columns = newMasterColumn
            # add the old DataFrame to the new Dataframe
            df = df.append(df_concat, ignore_index=True)
            print(df)
            # write the new DataFrame to the output file
            try:
                df.to_csv(outputFile, index=False)
                print('Standardized file successfully created...')
            except FileNotFoundError:
                print('ERROR: There was a problem creating the file...')
        except FileNotFoundError:
            print('File not found. Check file or directory...')
    else:
        print('%s does not exist...' % columnName)


# this function updates the record file every time the program is ran
def updateRecord():
    now = datetime.datetime.now()
    # creates a list for the record file
    recordsList = []
    # creates the first row, which be used as a template for the records
    recordsList.append(['City', 'Year', 'Standardized'])

    # searches through the raw data directories and files
    # loop that adds records to the recordsList
    for rawPath, rawName, rawFilename in os.walk('./raw_data/311_raw/'):
        for cities in rawName:
            city = cities
            # print(city)
            cityPath = os.path.join(rawPath, city)
            # print(cityPath)
            for filename in os.listdir(cityPath):
                rawFile = filename.split('.')
                try:
                    year = int(rawFile[0])
                except:
                    year = rawFile[0]
                # print(year)
                extension = rawFile[1]
                if isinstance(year, int):
                    if extension == 'csv':
                        print(city + " : " + filename)
                        standardizedFilePath = './standardized_data/311_standardized/' + \
                            city + '/' + str(year) + \
                            '_standardized.' + extension
                        print(standardizedFilePath)
                        if os.path.exists(standardizedFilePath):
                            if year == now.year:
                                recordsList.append([city, year, 'Incomplete'])
                            else:
                                recordsList.append([city, year, 'Yes'])
                        else:
                            recordsList.append([city, year, 'No'])
                    else:
                        print("File is not a csv file.")
                else:
                    print("Year is not integer.")
    # print(recordsList)

    # assigns the cells to the appropriate cell value
    # saves the record file
    for row in range(1, len(recordsList) + 1):
        for col in range(1, 4):
            record = recordsList[row - 1][col - 1]
            cell = recordsSheet.cell(column=col, row=row, value=record)
            # print(cell.value)
            records.save('Standardization_Records_test.xlsx')


# this function reads the record file and standardizes files based on it
def readRecordFile():
    for rows in recordsSheet.iter_rows(min_row=2):
        now = datetime.datetime.now()
        city = rows[0].value
        year = rows[1].value
        standardized = rows[2].value
        if standardized == 'No' or standardized == None or standardized == 'Incomplete':
            print(city + " " + str(year))
            print()
            for rawPath, rawName, rawFilename in os.walk('./raw_data/311_raw/'):
                for newRawName in rawName:
                    if newRawName == city:
                        print(newRawName)
                        print()
                        rawFile = str(year) + '.csv'
                        standardizedFile = str(year) + '_standardized.csv'
                        rawFilePath = os.path.join(rawPath, city, rawFile)
                        print(rawFilePath)
                        if os.path.exists(rawFilePath):
                            print('%s file exist.' % rawFilePath)
                            for standardizedPath, standardizedName, standardizedFilename in os.walk('./standardized_data/311_standardized/'):
                                standardizedFilePath = os.path.join(
                                    standardizedPath, city)
                                outputFilePath = os.path.join(
                                    standardizedPath, city, standardizedFile)
                                print(standardizedFilePath)
                                if os.path.isdir(standardizedFilePath):
                                    print('%s exist.' % standardizedFilePath)
                                    createFile(rawFilePath, city,
                                               outputFilePath)
                                    position = rows[2].coordinate
                                    if year == now.year:
                                        recordsSheet[position].value = 'Incomplete'
                                        records.save(
                                            'Standardization_Records_test.xlsx')
                                        break
                                    else:
                                        recordsSheet[position].value = 'Yes'
                                        records.save(
                                            'Standardization_Records_test.xlsx')
                                        break
                                else:
                                    print('%s does not exist.' %
                                          standardizedFilePath)
                                    try:
                                        os.makedirs(standardizedFilePath)
                                    except OSError:
                                        print(
                                            "Creation of the directory %s failed" % standardizedFilePath)
                                    else:
                                        print(
                                            "Successfully created the directory %s " % standardizedFilePath)
                                        createFile(
                                            rawFilePath, city, outputFilePath)
                                        position = rows[2].coordinate
                                        if year == now.year:
                                            recordsSheet[position].value = 'Incomplete'
                                            records.save(
                                                'Standardization_Records_test.xlsx')
                                            break
                                        else:
                                            recordsSheet[position].value = 'Yes'
                                            records.save(
                                                'Standardization_Records_test.xlsx')
                                            break
                        else:
                            print('%s file does not exist.' % rawFilePath)
        else:
            print(city + '.csv file is already standardized.')


# the program is ran here
print('+-------------------------+')
print('| STANDARDIZATION PROGRAM |')
print('+-------------------------+')
print('Version 2.0.3')
print('Elijah Toussaint')
print('Farzana Yusuf' + '\n')

# the Master_List excel file is read and used throughout the program
try:
    masterFile = 'Master_List_311_Cities_new (3).xlsx'
    wb = load_workbook(filename=masterFile)
    ws = wb.active
except FileNotFoundError:
    print('%s not found. Check file or directory...' % masterFile)

# the Standardization_Records excel file is read and used here
# if file does not exist, a record file is created
try:
    recordsFile = 'Standardization_Records_test.xlsx'
    records = load_workbook(filename=recordsFile)
    recordsSheet = records.active
except FileNotFoundError:
    print('%s not found...' % recordsFile)
    recordsFile = 'Standardization_Records_test.xlsx'
    records = openpyxl.Workbook()
    recordsSheet = records.active
    print('Creating %s file...' % recordsFile)
    records.save(recordsFile)

updateRecord()
readRecordFile()
updateRecord()